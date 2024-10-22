# Kade Hennacy 9/4/2024

# To best understand this script, read these comments in order starting at 1. Comment #2 is at the bottom.

# 1: Import necessary libraries.
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, Label, Button, Frame, IntVar, Checkbutton
from tkinter.ttk import Combobox, Spinbox
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font

# 3: Initialize global variables that will be used across functions.
df = None
file_path = ""

# 4: Define the main functions that handle file loading, processing, and saving.

# Function to load the input file when the "Load Input File" button is pressed.
def load_file():
    global file_path
    format_setting = format_combo.get()

    # 5: Set the acceptable file types based on the selected format setting.
    if format_setting == "Gmetrix Raw Data":
        file_type = [("CSV files", "*.csv")]
    else:
        file_type = [("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")]

    # 6: Open a file dialog for the user to select a file.
    file_path = filedialog.askopenfilename(filetypes=file_type)
    if file_path:
        filename = os.path.basename(file_path)
        file_label.config(text=f"Loaded file: {filename}", font=('Helvetica', 10, 'bold'))

# Function to save the processed file when the "Save Formatted File" button is pressed.
def save_file():
    # 7: Process the file based on the selected format setting.
    process_file()
    if 'df' not in globals() or df is None:
        messagebox.showerror("Error", "No processed data available. Please load and process a file first.")
        return

    # 8: Open a file dialog for the user to specify the output file path.
    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not output_file_path:
        return

    # 9: Create a new workbook and worksheet using openpyxl.
    wb = openpyxl.Workbook()
    ws = wb.active

    format_setting = format_combo.get()

    # 10: Write headers if the format setting requires them.
    if format_setting in ["Gmetrix for CTRL-R Import", "NFR Rise Up for CTRL-R Import", "NorthStar for CTRL-R Import"]:
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = header_alignment
        start_row = 2  # Data starts from the second row
    else:
        start_row = 1  # Data starts from the first row

    # 11: Write data to the worksheet.
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start_row):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 12: Apply general formatting if needed.
    if "for CTRL-R Import" not in format_setting:
        general_formatting(ws)

    # 13: Save the workbook to the specified output file path.
    wb.save(output_file_path)
    messagebox.showinfo("Success", f"Data processed and saved to {output_file_path}")

# Function to process the file based on the selected format setting.
def process_file():
    global df, file_path
    format_setting = format_combo.get()

    # 14: Read the input file into a pandas DataFrame.
    try:
        if format_setting in ["Gmetrix for CTRL-R Import", "NFR Rise Up for CTRL-R Import", "NorthStar for CTRL-R Import"]:
            df = pd.read_excel(file_path)
        else:
            if file_path.endswith('.csv'):
                sanitize_csv()
                df = pd.read_csv(file_path, header=None)
            else:
                df = pd.read_excel(file_path, header=None)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process the file\n{e}")
        return

    # 15: Call the appropriate processing function based on the format setting.
    if format_setting == "Gmetrix Raw Data":
        process_gmetrix()
    elif format_setting == "Gmetrix for CTRL-R Import":
        process_ctrlr_import()
    elif format_setting == "NFR Rise Up for CTRL-R Import":
        process_nfr_ctrlr_import()
    elif format_setting == "NorthStar for CTRL-R Import":
        process_northstar_ctrlr_import()

# 16: Define functions specific to each format setting.

# Function to process Gmetrix Raw Data.
def process_gmetrix():
    global df
    if df is None:
        messagebox.showerror("Error", "No file loaded. Please load a CSV file first.")
        return

    # 17: Remove columns containing "Minutes Spent" or "Score".
    columns_to_remove = []
    for col in df.columns:
        if df[col].apply(lambda x: str(x).strip().lower() in ["minutes spent", "score"]).any():
            columns_to_remove.append(col)
    df.drop(columns_to_remove, axis=1, inplace=True)

    # 18: Sort data if required.
    sort_order = sort_order_combo.get()
    if sort_order != "Unsorted":
        ascending_order = sort_order == "Ascending"
        post_assessment_cols = []

        for col in df.columns:
            if df[col].apply(lambda x: "Post-Assessment" in str(x)).any():
                post_assessment_cols.append(col)

        for post_assessment_col in post_assessment_cols:
            test_score_rows = df[df[post_assessment_col] == 'Test Score'].index

            for test_score_row in test_score_rows:
                next_blank_row = df[df.index > test_score_row][post_assessment_col].first_valid_index()
                end_row = df[df.index >= next_blank_row][post_assessment_col].isna().idxmax() if pd.notna(next_blank_row) else len(df)

                scores_data = df.iloc[test_score_row + 1:end_row].copy()
                scores_data[post_assessment_col] = scores_data[post_assessment_col].str.rstrip('%').apply(pd.to_numeric, errors='coerce')
                scores_data.dropna(subset=[post_assessment_col], inplace=True)
                sorted_data = scores_data.sort_values(by=post_assessment_col, ascending=ascending_order)
                sorted_data[post_assessment_col] = sorted_data[post_assessment_col].apply(lambda x: f"{x}%")
                df.iloc[test_score_row + 1:end_row] = sorted_data

# Function to process data for CTRL-R Import from Gmetrix.
def process_ctrlr_import():
    global df
    if df is None:
        messagebox.showerror("Error", "No file loaded. Please load an Excel file first.")
        return

    required_columns = ['Course Name', 'First Name', 'Last Name', 'Score']
    if not all(col in df.columns for col in required_columns):
        messagebox.showerror("Error", "Input file does not contain the required columns.")
        return

    df = df.dropna(subset=['Course Name', 'First Name', 'Last Name'])

    # 19: Combine 'First Name' and 'Last Name' to create 'Students' and 'Student Course Name'.
    df['Students'] = df['First Name'].astype(str).str.strip() + ' ' + df['Last Name'].astype(str).str.strip()
    df['Student Course Name'] = df['Students'] + ' - ' + df['Course Name'].astype(str).str.strip()

    df['Score'] = df['Score'].astype(str).str.rstrip('%').astype(float)

    passing_percentage = passing_percentage_var.get()
    df['Status'] = df['Score'].apply(lambda x: 'Complete' if x >= passing_percentage else 'In Progress')
    df['Exam Score'] = df['Score']
    df['Certificates Earned'] = ''
    df['Course Completion Date'] = ''

    # 20: Reorder columns for the output.
    df_output = df[['Students', 'Course Name', 'Status', 'Exam Score',
                    'Certificates Earned', 'Course Completion Date', 'Student Course Name']]
    df = df_output

# Function to process data for CTRL-R Import from NFR Rise Up.
def process_nfr_ctrlr_import():
    global df
    if df is None:
        messagebox.showerror("Error", "No file loaded. Please load an Excel file first.")
        return

    required_columns = ['FIRST NAME', 'LAST NAME', 'COURSE/EXAM', 'TYPE', 'STATUS', 'COMPLETED']
    if not all(col in df.columns for col in required_columns):
        messagebox.showerror("Error", "Input file does not contain the required columns.")
        return

    # 21: Filter and process data.
    df = df[df['TYPE'].isin(['Exam', 'Exam Retest'])]
    df['Students'] = df['FIRST NAME'].astype(str).str.strip() + ' ' + df['LAST NAME'].astype(str).str.strip()
    df['Student Course Name'] = df['Students'] + ' - ' + df['COURSE/EXAM'].astype(str).str.strip()
    df['Status'] = df['STATUS'].apply(lambda x: 'Complete' if x.upper() == 'PASSED' else 'In Progress')
    df['Exam Score'] = df['STATUS'].apply(lambda x: 'PASS' if x.upper() == 'PASSED' else 'FAIL')
    df['Course Name'] = df['COURSE/EXAM']
    df['Course Completion Date'] = df['COMPLETED']
    df['Certificates Earned'] = ''

    df_output = df[['Students', 'Course Name', 'Status', 'Exam Score',
                    'Certificates Earned', 'Course Completion Date', 'Student Course Name']]
    df = df_output

# Function to process data for CTRL-R Import from NorthStar.
def process_northstar_ctrlr_import():
    global df
    if df is None:
        messagebox.showerror("Error", "No file loaded. Please load an Excel file first.")
        return

    required_columns = ['First Name', 'Last Name']
    if not all(col in df.columns for col in required_columns):
        messagebox.showerror("Error", "Input file does not contain 'First Name' and 'Last Name' columns.")
        return

    # 22: Combine 'First Name' and 'Last Name' to create 'Students' and 'Student Course Name'.
    df['Students'] = df['First Name'].astype(str).str.strip() + ' ' + df['Last Name'].astype(str).str.strip()
    df['Course Name'] = 'Northstar Digital Literacy'
    df['Student Course Name'] = df['Students'] + ' - ' + df['Course Name']

    # 23: Identify 'Certificate Earned' columns and count total certificates.
    certificate_columns = [col for col in df.columns if 'Certificate Earned' in col]
    if not certificate_columns:
        messagebox.showerror("Error", "No 'Certificate Earned' columns found in the input file.")
        return

    df['Total Certificates'] = df[certificate_columns].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
    df['Certificates Earned'] = df['Total Certificates'].astype(int)

    # 24: Set 'Exam Score' and 'Status' based on total certificates.
    passing_certificates = northstar_passing_certificates_var.get()
    df['Exam Score'] = df['Total Certificates'].apply(lambda x: 'Passed' if x >= passing_certificates else 'Failed')
    df['Status'] = df['Exam Score'].apply(lambda x: 'Complete' if x == 'Passed' else 'In Progress')
    df['Course Completion Date'] = ''

    df_output = df[['Students', 'Course Name', 'Status', 'Exam Score',
                    'Certificates Earned', 'Course Completion Date', 'Student Course Name']]
    df = df_output

# Function to sanitize CSV files by ensuring equal number of commas in each row.
def sanitize_csv():
    global file_path
    with open(file_path, 'r') as file:
        lines = file.readlines()

    max_commas = max(line.count(',') for line in lines)

    adjusted_lines = []
    for line in lines:
        current_commas = line.count(',')
        if current_commas < max_commas:
            line = line.strip('\n') + ',' * (max_commas - current_commas) + '\n'
        adjusted_lines.append(line)

    with open(file_path, 'w') as file:
        file.writelines(adjusted_lines)

# Function to apply general formatting to the worksheet.
def general_formatting(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            # 25: Apply word wrap if selected.
            if word_wrap_var.get() == 1:
                cell.alignment = Alignment(wrap_text=True)

            # 26: Apply text centering if selected.
            if center_text_var.get() == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=cell.alignment.wrap_text)

            # 27: Calculate maximum length for autosizing columns.
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # 28: Autosize or manually resize columns based on user selection.
        if autosize_col_var.get() == 1:
            ws.column_dimensions[column].width = max_length + 2
        elif resize_col_var.get() == 1:
            ws.column_dimensions[column].width = column_width_var.get()

# 29: Set up the GUI using tkinter.

# Initialize the main window.
root = tk.Tk()
root.title("Spreadsheet Formatter")
root.geometry("900x400")

# Create frames for organizing widgets.
frame_top = Frame(root)
frame_top.pack(pady=10)
frame_sort = Frame(root)
frame_sort.pack(pady=(10, 0))
frame_bottom = Frame(root)
frame_bottom.pack(pady=0)

# 30: Add labels and instructions.
file_label = Label(frame_bottom, wraplength=600, justify="left")
file_label.pack(pady=10)

instruction_text = "This program loads a spreadsheet and formats it according to the format setting."
label_instructions = Label(frame_bottom, text=instruction_text, wraplength=600, justify="left")
label_instructions.pack(pady=(30, 15))

additional_instruction_label = Label(frame_bottom, wraplength=600, justify="left")
additional_instruction_label.pack(pady=10)

# 31: Create the format setting dropdown.
format_setting_label = Label(frame_top, text="Format Setting")
format_setting_label.pack(side=tk.LEFT, padx=10)

format_combo = Combobox(frame_top, state="readonly", width=25)
format_combo['values'] = ("Gmetrix Raw Data", "Gmetrix for CTRL-R Import",
                          "NFR Rise Up for CTRL-R Import", "NorthStar for CTRL-R Import", "General Formatting")
format_combo.current(1)
format_combo.pack(side=tk.LEFT, padx=10)

# 32: Create the sort order dropdown (only visible for certain settings).
sort_order_label = Label(frame_sort, text="Sort Order")
sort_order_combo = Combobox(frame_sort, state="readonly", width=15)
sort_order_combo['values'] = ("Ascending", "Descending", "Unsorted")
sort_order_combo.current(1)

# 33: Create buttons for loading and saving files.
load_button = Button(frame_top, text="Load Input File", command=load_file)
load_button.pack(side=tk.LEFT, padx=10)

save_button = Button(frame_top, text="Save Formatted File", command=save_file)
save_button.pack(side=tk.LEFT, padx=10)

# 34: Define functions to handle UI updates based on user interactions.

# Function to update instructions and UI elements when the format setting changes.
def update_instruction(event=None):
    format_setting = format_combo.get()
    if format_setting == "Gmetrix Raw Data":
        additional_instruction_label.config(text="Gmetrix Raw Data - This setting takes a CSV Gmetrix student progress report as input. It removes the 'Minutes Spent' and 'Score' columns, sizes columns to 18, and centers and wraps all text, and outputs an XLSX file. To sort the rows by post-assessment score, select an option from the Sort Order dropdown. Descending places the highest scores at the top of the sheet.")
        sort_order_label.grid(row=0, column=0, padx=(10, 2), sticky='e')
        sort_order_combo.grid(row=0, column=1, padx=(2, 10), sticky='w')
        # Show formatting options.
        show_formatting_options()
        passing_percentage_label.grid_remove()
        passing_percentage_spin.grid_remove()
        northstar_passing_certificates_label.grid_remove()
        northstar_passing_certificates_spin.grid_remove()
    elif format_setting == "General Formatting":
        additional_instruction_label.config(text="General Formatting - This setting takes any CSV or XLSX file and sizes columns to 18 and centers and wraps text.")
        sort_order_label.grid_remove()
        sort_order_combo.grid_remove()
        # Show formatting options.
        show_formatting_options()
        passing_percentage_label.grid_remove()
        passing_percentage_spin.grid_remove()
        northstar_passing_certificates_label.grid_remove()
        northstar_passing_certificates_spin.grid_remove()
    elif format_setting == "Gmetrix for CTRL-R Import":
        additional_instruction_label.config(text="Gmetrix for CTRL-R Import - This setting takes a CSV Gmetrix student progress report as input. It removes and combines columns to create a file compatible with the import feature on the CTRL-R All Student Grades report.")
        hide_all_formatting_options()
        passing_percentage_label.grid(row=0, column=0, padx=(10, 2), sticky='e')
        passing_percentage_spin.grid(row=0, column=1, padx=(2, 10), sticky='w')
    elif format_setting == "NFR Rise Up for CTRL-R Import":
        additional_instruction_label.config(text="NFR Rise Up for CTRL-R Import - This setting takes an Excel file exported from the NFR Rise Up platform and formats it for CTRL-R import. It processes Exam and Exam Retest entries, combines columns, and creates a compatible file.")
        hide_all_formatting_options()
    elif format_setting == "NorthStar for CTRL-R Import":
        additional_instruction_label.config(text="NorthStar for CTRL-R Import - This setting processes a NorthStar exported Excel file. It counts the total 'Certificate Earned' columns per student, sets the 'Exam Score' to 'Passed' if the student has earned the specified number of certificates, and updates the 'Status' accordingly.")
        hide_all_formatting_options()
        northstar_passing_certificates_label.grid(row=0, column=0, padx=(10, 2), sticky='e')
        northstar_passing_certificates_spin.grid(row=0, column=1, padx=(2, 10), sticky='w')

# Function to show formatting options.
def show_formatting_options():
    word_wrap_check.grid(row=0, column=2, padx=(10, 2), sticky='w')
    center_text_check.grid(row=0, column=3, padx=(2, 2), sticky='w')
    autosize_col_check.grid(row=0, column=4, padx=(2, 2), sticky='w')
    resize_col_check.grid(row=0, column=5, padx=(2, 2), sticky='w')
    if resize_col_var.get() == 1:
        column_width_spin.grid(row=0, column=6, padx=(2, 2), sticky='w')
        px_label.grid(row=0, column=7, sticky='w')
    else:
        column_width_spin.grid_remove()
        px_label.grid_remove()

# Function to hide all formatting options.
def hide_all_formatting_options():
    sort_order_label.grid_remove()
    sort_order_combo.grid_remove()
    word_wrap_check.grid_remove()
    center_text_check.grid_remove()
    autosize_col_check.grid_remove()
    resize_col_check.grid_remove()
    column_width_spin.grid_remove()
    px_label.grid_remove()
    passing_percentage_label.grid_remove()
    passing_percentage_spin.grid_remove()
    northstar_passing_certificates_label.grid_remove()
    northstar_passing_certificates_spin.grid_remove()

# Function to handle the 'Resize Columns' checkbutton.
def handle_resize_checkbutton():
    if resize_col_var.get() == 1:
        autosize_col_var.set(0)
        column_width_spin.grid()
        px_label.grid()
    else:
        column_width_spin.grid_remove()
        px_label.grid_remove()

# Function to handle the 'Autosize Columns' checkbutton.
def handle_autosize_checkbutton():
    if autosize_col_var.get() == 1:
        resize_col_var.set(0)
        column_width_spin.grid_remove()
        px_label.grid_remove()

# 35: Initialize variables for UI inputs.
word_wrap_var = IntVar(value=1)
center_text_var = IntVar(value=1)
autosize_col_var = IntVar(value=1)
resize_col_var = IntVar(value=0)
column_width_var = IntVar(value=18)
passing_percentage_var = IntVar(value=70)
northstar_passing_certificates_var = IntVar(value=5)

# 36: Create UI elements for formatting options.
word_wrap_check = Checkbutton(frame_sort, text="Word Wrap", variable=word_wrap_var)
center_text_check = Checkbutton(frame_sort, text="Center Text", variable=center_text_var)
resize_col_check = Checkbutton(frame_sort, text="Resize Columns", variable=resize_col_var, command=handle_resize_checkbutton)
autosize_col_check = Checkbutton(frame_sort, text="Autosize Columns", variable=autosize_col_var, command=handle_autosize_checkbutton)
column_width_spin = Spinbox(frame_sort, from_=10, to=50, textvariable=column_width_var, width=5)
px_label = Label(frame_sort, text="Points")
passing_percentage_label = Label(frame_sort, text="Passing Percentage")
passing_percentage_spin = Spinbox(frame_sort, from_=0, to=100, textvariable=passing_percentage_var, width=5)

# New UI elements for NorthStar certificates needed to pass
northstar_passing_certificates_label = Label(frame_sort, text="Certificates Needed to Pass")
northstar_passing_certificates_spin = Spinbox(frame_sort, from_=0, to=100, textvariable=northstar_passing_certificates_var, width=5)

# 37: Position the formatting options in the grid.
word_wrap_check.grid(row=0, column=2, padx=(10, 2), sticky='w')
center_text_check.grid(row=0, column=3, padx=(2, 2), sticky='w')
autosize_col_check.grid(row=0, column=4, padx=(2, 2), sticky='w')
resize_col_check.grid(row=0, column=5, padx=(2, 2), sticky='w')
column_width_spin.grid(row=0, column=6, padx=(2, 2), sticky='w')
px_label.grid(row=0, column=7, sticky='w')

# 38: Bind events and initialize UI.
format_combo.bind("<<ComboboxSelected>>", update_instruction)
update_instruction()
handle_resize_checkbutton()

# 2: Start the event loop of the user interface.
root.mainloop()
